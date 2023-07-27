import openpyxl
import datetime
import warnings
warnings.filterwarnings('ignore')

colors = ['Blue','Red','Brown','Green','Orange','Purple']
subTaskKeys = ['name','startDate','endDate','color']

def getExportFileNameAndOpenWorkbook ():
    excelFileName = input("Enter the file name of the excel workbook exported for the TFC Project Planner: \n")

    projectPlannerExportObj = openpyxl.load_workbook(excelFileName)
    plannerWS = projectPlannerExportObj["Project tasks"]
    getNeededDataFromExcel(plannerWS)

def getNeededDataFromExcel(plannerExportWS):
    global majorTaskNameAndNumber
    majorTaskNameAndNumber = {}

    getMajorTaskNames(plannerExportWS)
   
    global subTasks
    subTasks = {}

    colorCount = 0
    for key in majorTaskNameAndNumber:
        subTasks[key] = getSubTasks(key, plannerExportWS, colorCount)
        colorCount += 1


def getMajorTaskNames(projectTasksWS):

    count = 10
    cell = projectTasksWS['B'+str(count)]
    while (cell.value is not None):
        if ('.' not in cell.value):
           tNum = str(cell.value)
           tName = projectTasksWS['C'+str(count)].value
           majorTaskNameAndNumber[tNum] = tName
        count += 1
        cell = projectTasksWS['B'+str(count)]

def getSubTasks(taskNum, exportWS, colCount):
    allSubTaskInfo = []

    count = 10
    val = exportWS['B'+str(count)].value

    while(val is not None):
        t = str(val)
        if (t.startswith(str(taskNum))) and (t.count('.') == 1):
            subTaskInfo = {}
            subTaskInfo['name'] = exportWS['C'+str(count)].value
            tStart = exportWS['D'+str(count)].value
            subTaskInfo['startDate'] = datetime.date(tStart.year, tStart.month, tStart.day)
            tEnd = exportWS['E'+str(count)].value
            subTaskInfo['endDate'] = datetime.date(tEnd.year, tEnd.month, tEnd.day)
            subTaskInfo['color'] = colors[colCount]
            allSubTaskInfo.append(subTaskInfo)
        count += 1
        val = exportWS['B'+str(count)].value

    return allSubTaskInfo

def openTimelineTemplate():
    docPath = 'TFC Project-Timeline Code Export '+str(datetime.date.today())+'.xlsx'
    wb = openpyxl.load_workbook("SCRUBBED TFC Timeline TABLE.xlsx")
    templateWS = wb['ProjectTimeline']
    printData(templateWS)
    wb.save(docPath)
    wb.close()


def printData(templateWS):

    rowNum = 6
    insertAbove = 7
    for key, v in majorTaskNameAndNumber.items():

        categoryCell = templateWS['A'+str(rowNum)]
        categoryCell.value = v
        tasks = subTasks[key]
        for t in tasks:
            taskCell = templateWS['B'+str(rowNum)]
            taskCell.value = t['name']

            startCell = templateWS['C'+str(rowNum)]
            startCell.value = t['startDate']

            endCell = templateWS['D'+str(rowNum)]
            endCell.value = t['endDate']

            colorCell = templateWS['E'+str(rowNum)]
            colorCell.value = t['color']

            t = templateWS.tables['Table13']
            t.ref = 'A4:L'+str(rowNum)

            templateWS.insert_rows(insertAbove)
            rowNum+= 1
            insertAbove+=1

getExportFileNameAndOpenWorkbook()
openTimelineTemplate()
